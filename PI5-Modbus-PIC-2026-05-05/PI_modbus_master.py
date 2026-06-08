#!/home/alphotronic/Desktop/ALPHO/PI5-Modbus-PIC-2026-05-05/venv/bin/python3

"""-----------------------------------------------------------------------------------------
PI5 Modbus/RTU Master - Reading data registers from PIC MF40 controller
via RS-485 HAT (RB-RS485) on Raspberry Pi 5

PIC MF40 Register Map (Holding Registers, Function Code 3):
---------------------------------------------------------------------------
RW Registers (address 0-23, each entry = 2 registers = 4 bytes = float32):
  Addr 0,1   : SCADA_Cmd0 (CMD32)
  Addr 2,3   : P_SET      (Power Setpoint)
  Addr 4,5   : TEMP_SET   (Temperature Setpoint)
  Addr 6,7   : TIMER_SET  (Timer Setpoint)
  Addr 8,9   : PID_Kp
  Addr 10,11 : PID_Ki
  Addr 12,13 : PID_Kd
  Addr 14,15 : GenOffUse   (2 bytes, uint16 stored as float)
  Addr 16,17 : StopAtTemp  (2 bytes, uint16 stored as float)
  Addr 18,19 : TimerDelay  (2 bytes, uint16 stored as float)
  Addr 20,21 : Spare
  Addr 22,23 : Spare

RO Registers (address 200-215, each entry = 2 registers = 4 bytes = float32):
  Addr 200,201 : SCADA_Status0 / STATUS32
  Addr 202,203 : P       (Actual Power)
  Addr 204,205 : TEMP    (Actual Temperature)
  Addr 206,207 : TIMER   (Actual Timer)
  Addr 208,209 : I       (Actual Current)
  Addr 210,211 : F       (Actual Frequency)
  Addr 212,213 : COS_PHI (Actual Cos-Phi)
  Addr 214,215 : REG%    (Regulator %)

Note: Addresses 1000+ enable little-endian byte order on PIC side.
--------------------------------------------------------------------------------------------
"""

import csv
import glob
import os
import serial
import struct
import time
from datetime import datetime, timezone

from pymodbus.client import ModbusSerialClient
from pymodbus.exceptions import ModbusException

from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side, NamedStyle, colors, PatternFill
from openpyxl.chart import ScatterChart, Reference, Series

# ============================================================================
# Bit definitions for status registers
# ============================================================================
BIT0      = 0
BIT1      = 1

# ============================================================================
# Configuration - adjust these to match your PIC MF40 setup
# ============================================================================
SERIAL_PORT  = '/dev/ttyAMA0'       # RS-485 HAT on Pi 5 (/dev/serial0 -> /dev/ttyAMA0), RP1 chip not CPU(BCM2712)
BAUDRATE     = 57600                # Default baud rate for PIC MF40 : 57600
SLAVE_ID     = 50                   # PIC Modbus slave address (MF40_Userdata.MB_Slave_Adr)
BYTESIZE     = 8
PARITY       = 'N'                  # No parity
STOPBITS     = 1
TIMEOUT      = 3                    # seconds

# ============================================================================
# Register definitions
# ============================================================================

# RW registers (function code 3 CMD sent to PIC32(SLAVE ID 50), starting at address 2)
RW_REGISTERS = {
    2:  "P_SET",
    4:  "TEMP_SET",
    6:  "TIMER_SET",
    8:  "PID_Kp",
    10: "PID_Ki",
    12: "PID_Kd",
    14: "GenOffUse",
    16: "StopAtTemp",
    18: "TimerDelay",
}

# RO registers (function code 3 CMD sent to PIC32(SLAVE ID 50), starting at address 200)
RO_REGISTERS = {
    202: "P",
    204: "TEMP",
    206: "TIMER",
    208: "I",
    210: "FREQ",
    212: "COS_PHI",
    214: "REG%",
}

def plot_graph(sheet, num_rows, psn_string=""):
    """Plot Set temperature (Blue) and Measured temperature (Red) vs Pair number
    as a scatter chart on the right side of the same sheet.

    Args:
        sheet: Active openpyxl worksheet (data already written)
        num_rows: Number of data rows (excluding header)
        psn_string: Production Serial Number string for chart title
    """
    chart = ScatterChart()
    chart.title  = f"[PSN : {psn_string}] Temperature/°C vs Pair Number"
    chart.style  = 13                        # clean built-in style
    chart.width  = 30                        # cm – wide enough for 1800 pairs
    chart.height = 15

    # --- X-axis scaling (Pair Number) ---
    #chart.x_axis.title       = "Pair Number"
    chart.x_axis.delete      = False         # force axis labels visible
    chart.x_axis.tickLblPos  = "low"         # show tick labels at bottom
    chart.x_axis.numFmt      = "0"           # integer format
    chart.x_axis.majorGridlines = None       # no vertical gridlines (cleaner)

    # --- Y-axis scaling (Temperature °C) ---
    # Note: axis title omitted — chart title already says "Temperature"
    chart.y_axis.delete      = False         # force axis labels visible
    chart.y_axis.tickLblPos  = "low"         # show tick labels at left
    chart.y_axis.numFmt      = "0"           # integer format

    # X-axis data: column B (Pair number), from row 2 to num_rows+1
    x_values = Reference(sheet, min_col=2, min_row=2, max_row=num_rows + 1)

    # Series 1 – Set temperature (column C) in BLUE
    set_values = Reference(sheet, min_col=3, min_row=2, max_row=num_rows + 1)
    series_set = Series(set_values, x_values, title="Set temperature")
    series_set.graphicalProperties.line.solidFill = "0000FF"
    series_set.graphicalProperties.line.width     = 40000   # EMU (~1.0 pt)
    series_set.marker.symbol = "none"                       # line only, no markers
    chart.series.append(series_set)

    # Series 2 – Measured temperature (column D) in RED
    meas_values = Reference(sheet, min_col=4, min_row=2, max_row=num_rows + 1)
    series_meas = Series(meas_values, x_values, title="Measured temperature")
    series_meas.graphicalProperties.line.solidFill = "FF0000"
    series_meas.graphicalProperties.line.width     = 40000
    series_meas.marker.symbol = "none"
    chart.series.append(series_meas)

    # Place chart to the right of the data (column G, row 1)
    sheet.add_chart(chart, "G1")


def create_record_temperatures_excel_table(script_dir, timestamp, paired_rows, psn_string="", error_string="NO ERROR"):

    global workbook

    sheet                 = workbook.active
    header_data           = ('Date/time-stamp', 'Pair', 'Set temperature', 'Measured temperature', 'ERROR')
    cell_range            = sheet['A1':'E1']
    row_a                 = sheet[1]
    #apply filter in active sheet
    sheet.auto_filter.ref = "A1:E1"
    # Creating a few styles for the header data inside cell.
    header                = NamedStyle(name = "header")
    header.font           = Font(bold=True, size = 12)
    header.border         = Border(bottom=Side(border_style="thick"))
    header.alignment      = Alignment(horizontal="center", vertical="center")

    #run loop for header data
    header_data_counter   = 0
    # working with individual cells
    for dataCell in row_a:
        dataCell.style       = header
        dataCell.value       = header_data[header_data_counter]
        if dataCell.value    == 'Set temperature': 
            dataCell.font    = Font(bold=True, size = 15, color = "0000FF")
        elif dataCell.value  == 'Measured temperature':
            dataCell.font    = Font(bold=True, size = 15, color = "FF0000")
        header_data_counter  +=1

    # Write paired data rows starting from row 2
    # paired_rows format: (utc_str, pair_num, set_dec, meas_dec)
    for row_idx, row_data in enumerate(paired_rows, start=2):
       
        sheet.cell(row=row_idx, column=1, value=row_data[0])  # A: Date/time-stamp
        sheet.cell(row=row_idx, column=2, value=row_data[1])  # B: Pair number
        sheet.cell(row=row_idx, column=3, value=row_data[2])  # C: Set temperature
        sheet.cell(row=row_idx, column=4, value=row_data[3])  # D: Measured temperature
        
        error_cell = sheet.cell(row=row_idx, column=5, value=error_string)   # E: Error
        # if ERROR then fill the cell with RED colour
        if error_string != "NO ERROR":
            error_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            error_cell.font = Font(color="FFFFFF", bold=True) # Optional: make text white and bold for better readability on red

    # Extend auto-filter to cover all data rows
    sheet.auto_filter.ref = f"A1:E{len(paired_rows) + 1}"

    # Plot temperature graph on the right side of the same sheet
    plot_graph(sheet, len(paired_rows), psn_string)

    #Save the excel file with same naming nomenclature as CSV (TL_data_{timestamp}.xlsx)
    excel_path = os.path.join(script_dir, f"TL_data_{timestamp}.xlsx")
    workbook.save(filename = excel_path)
    print(f"  [TL] Excel saved → {excel_path}  ({len(paired_rows)} paired rows)")

    

def registers_to_float(reg_hi, reg_lo):
    """Convert two 16-bit Modbus registers to a 32-bit float (big-endian/word swap).
    
    PIC uses cswap4s with big-endian byte order: high register first, then low register.

    >(Greater than sign) = This tells Python to use BIG_ENDIAN format. This means the most significant bytes come first
    H                    = Unsigned short 16-bits
    HH                   = expects two 16-bits = 32 bits integer
    """
    raw_bytes = struct.pack('>HH', reg_hi, reg_lo)
    return struct.unpack('>f', raw_bytes)[0]


def read_holding_register_float(client, address, slave_id=SLAVE_ID):
    """Read a 32-bit float value from two consecutive holding registers.
    
    Args:
        client: Modbus client
        address: Starting register address 
        slave_id: Modbus slave address
    
    Returns:
        float value or None on error
    """
    try:
        result = client.read_holding_registers(address=address, count=2, device_id=slave_id)
        if result.isError():
            print(f"  Error reading address {address}: {result}")
            return None
        return registers_to_float(result.registers[0], result.registers[1])
    except ModbusException as e:
        print(f"  Modbus exception at address {address}: {e}")
        return None


def read_status_registers(client):
    """Read SCADA status registers."""
    print("\n--- Status Registers ---")
    
    temp_measurement_err = 0
    err_I = err_FI = err_PHI = err_Id = err_T = err_Uc = err_EXT = err_H2O = 0
    
    # SCADA_Cmd0 at address 0 (uint16)
    '''try:
        result = client.read_holding_registers(address=0, count=1, device_id=SLAVE_ID)
        if not result.isError():
            print(f"  [  0] SCADA_Cmd0               = {result.registers[0]} (0x{result.registers[0]:04X})")
    except ModbusException as e:
        print(f"  Error reading SCADA_Cmd0: {e}")'''

    # TEMP_SET at address 4 (float32)
    temp_set = read_holding_register_float(client, 4)
    if temp_set is not None:
        print(f"  [  4] TEMP_SET                 = {temp_set:.2f}")

    # P at address 2 (float32)
    p = read_holding_register_float(client, 2)
    if p is not None:
        print(f"  [2] P                      = {p:.2f}")

    # STATUS0 at address 200 (uint16)
    try:
        STATUS0_Temperature_Measurement = client.read_holding_registers(address = 200, count = 1, device_id = SLAVE_ID)
        if not STATUS0_Temperature_Measurement.isError():
            status0_val = STATUS0_Temperature_Measurement.registers[0]
            temp_measurement_err = (status0_val >> 13) & 1
            print(f"  [200] STATUS0                  = {status0_val} (0x{status0_val:04X})")
            print(f"  [200] Temp Meas Error (BIT13)  = {temp_measurement_err}")
    except ModbusException as e:
        print(f"  Error reading STATUS0: {e}")

    # STATUS1 at address 201 (uint16)
    try:
        STATUS1_LED = client.read_holding_registers(address = 201, count = 1, device_id = SLAVE_ID)
        if not STATUS1_LED.isError():
            status1_val = STATUS1_LED.registers[0]
            print(f"  [201] STATUS1                  = {status1_val} (0x{status1_val:04X})")
            
            err_I   = (status1_val >> 8) & 1
            err_FI  = (status1_val >> 9) & 1
            err_PHI = (status1_val >> 10) & 1
            err_Id  = (status1_val >> 11) & 1
            err_T   = (status1_val >> 12) & 1
            err_H2O = (status1_val >> 13) & 1
            err_Uc  = (status1_val >> 14) & 1
            err_EXT = (status1_val >> 15) & 1
            
            print(f"  [201] Error I (BIT8)           = {err_I}")
            print(f"  [201] Error FI (BIT9)          = {err_FI}")
            print(f"  [201] Error PHI (BIT10)        = {err_PHI}")
            print(f"  [201] Error Id (BIT11)         = {err_Id}")
            print(f"  [201] Error T (BIT12)          = {err_T}")
            print(f"  [201] Error H2O (BIT13)        = {err_H2O}")
            print(f"  [201] Error Uc (BIT14)         = {err_Uc}")
            print(f"  [201] Error EXT (BIT15)        = {err_EXT}")
    except ModbusException as e:
        print(f" Error reading STATUS1: {e}")    

    # TEMP at address 204 (float32)
    temp = read_holding_register_float(client, 204)
    if temp is not None:
        print(f"  [204] TEMP                     = {temp:.2f}")
    
    # TL_Busy at address 300 (uint16)
    try:
        TL_Busy = client.read_holding_registers(address=300, count=1, device_id=SLAVE_ID)
        if not TL_Busy.isError():
            print(f"  [300] TL_Busy                  = {TL_Busy.registers[0]} (0x{TL_Busy.registers[0]:04X})")
    except ModbusException as e:
        print(f"  Error reading TL_Busy: {e}")

    # TL_ready at address 301 (uint16)
    try:
        TL_ready = client.read_holding_registers(address=301, count=1, device_id=SLAVE_ID)
        if not TL_ready.isError():
            print(f"  [301] TL_ready                 = {TL_ready.registers[0]} (0x{TL_ready.registers[0]:04X})")
    except ModbusException as e:
        print(f"  Error reading TL_ready: {e}")

    '''
            Reading Temperature List registers
            BIG ENDIAN = High Register First, then Low Register
            10000 to 13599 = 3600 registers
            Read only if TL_ready == 1
    '''
    # Extracting bit from TL_BUSY and TL_READY registers
    val_busy  = TL_Busy.registers[0]  if not TL_Busy.isError()  else 0
    val_ready = TL_ready.registers[0] if not TL_ready.isError() else 0

    val_busy_ready = (val_busy << BIT1) | (val_ready << BIT0)
    print(f"  [V] TL_Busy_Ready            = {val_busy_ready} (0x{val_busy_ready:04X})")

    # Initialize static-like variable for status if it doesn't exist
    if not hasattr(read_status_registers, "val_busy_ready_STATUS"):
        read_status_registers.val_busy_ready_STATUS = 0

    if val_busy_ready == 1:
        read_status_registers.val_busy_ready_STATUS = 1

    if read_status_registers.val_busy_ready_STATUS == 1:
        '''
                Reading Production Serial Number registers
                BIG ENDIAN = High Register First, then Low Register
                308 to 323 = 16 registers
                Read only if TL_ready == 1
        '''
        PSN_START  = 308
        PSN_END    = 323

        TL_START   = 10000
        TL_END     = 13599
        TL_COUNT   = TL_END - TL_START + 1   # 3600 registers
        CHUNK_SIZE = 125                      # pymodbus max per request

        print(f"\n  [TL] TL_Ready detected — reading {TL_COUNT} registers ({TL_START}–{TL_END})...")

        tl_rows = []   # list of (address, value_dec, value_hex)
        read_errors = 0

        #--------------------------------------------------------------------
        # Read all 16 PSN registers (308–323) in one request
        #--------------------------------------------------------------------
        psn_string = ""
        try:
            psn_result = client.read_holding_registers(
                address=PSN_START, count=(PSN_END - PSN_START + 1), device_id=SLAVE_ID
            )
            if not psn_result.isError():
                psn_chars = []
                done = False
                for psn_result_val in psn_result.registers:
                    # Each register holds 2 ASCII bytes: high byte first
                    hi_byte = (psn_result_val >> 8) & 0xFF
                    lo_byte =  psn_result_val       & 0xFF
                    for psn_result_val_byte in (hi_byte, lo_byte):
                        if psn_result_val_byte == 0:               # null terminator → stop
                            done = True
                            break
                        if 0x20 <= psn_result_val_byte <= 0x7E:    # printable ASCII only
                            psn_chars.append(chr(psn_result_val_byte))
                    if done:
                        break
                psn_string = "".join(psn_chars)
                print(f"  [PSN] Production Serial Number = {psn_string!r}")
            else:
                psn_string = "READ_ERROR"
                print(f"  [PSN] Error reading PSN registers: {psn_result}")
        except ModbusException as e:
            psn_string = "MODBUS EXCEPTION OCCURRED"
            print(f"  [PSN] Modbus exception reading PSN: {e}")

        # ----------------------------------------------------------------
        # read all 3600 TL registers (10000 – 13599) and write CSV
        # ----------------------------------------------------------------
        addr = TL_START
        while addr <= TL_END:

            chunk = min(CHUNK_SIZE, TL_END - addr + 1)

            try:
                result = client.read_holding_registers(addr, count=chunk, device_id=SLAVE_ID)
                if not result.isError():
                    for i, reg_val in enumerate(result.registers):
                        tl_rows.append((addr + i, reg_val, f"0x{reg_val:04X}"))
                else:
                    print(f"  [TL] Error reading chunk at {addr}: {result}")
                    for i in range(chunk):
                        tl_rows.append((addr + i, "ERR", "ERR"))
                    read_errors += chunk

            except ModbusException as e:
                print(f"  [TL] Modbus exception at {addr}: {e}")
                for i in range(chunk):
                    tl_rows.append((addr + i, "ERR", "ERR"))
                read_errors += chunk

            addr += chunk

        # ----------------------------------------------------------------
        # Pair up raw register data: even offset = Set, odd offset = Measured
        # Pattern: addr 10000=Set, 10001=Measured, 10002=Set, 10003=Measured ...
        # ----------------------------------------------------------------
        
        utc_now = datetime.now(timezone.utc)
        timestamp = utc_now.strftime("%Y%m%d_%H%M%S")

        paired_rows = []
        # range(start, stop, step)
        for idx in range(0, len(tl_rows), 2):
            set_row = tl_rows[idx]          # even offset: set temperature
            if idx + 1 < len(tl_rows):
                meas_row = tl_rows[idx + 1]  # odd offset: measured temperature
            else:
                meas_row = ("N/A", "N/A", "N/A")  # safety: if odd total count
            # Each paired row: (row_utc_str, pair_number, set_dec, meas_dec)
            # pair_num = 1, 2, 3, 4, 5 till 1800
            pair_num = idx // 2 + 1
            # Generate a new timestamp with milliseconds for each row
            # [:-3] removes 3 digits from the 6 digits of microseconds and coverts to  milliseconds
            row_utc_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
            paired_rows.append((row_utc_str, pair_num, set_row[1], meas_row[1]))

        # create list of error as active_errors from the error flags 
        active_errors = []
        if temp_measurement_err: active_errors.append("Temp Measurement") # from STATUS0
        if err_I: active_errors.append("I")                              # from STATUS1
        if err_FI: active_errors.append("FI")                            # from STATUS1
        if err_PHI: active_errors.append("PHI")                          # from STATUS1
        if err_Id: active_errors.append("Id")                            # from STATUS1     
        if err_T: active_errors.append("T")                              # from STATUS1
        if err_H2O: active_errors.append("H2O")                          # from STATUS1
        if err_Uc: active_errors.append("Uc")                            # from STATUS1
        if err_EXT: active_errors.append("EXT")                          # from STATUS1
        
        error_string = ", ".join(active_errors) if active_errors else "NO ERROR"

        # Always write to a fresh timestamped CSV; remove the previous one first
        # This is the directory of a script
        script_dir    = os.path.dirname(os.path.abspath(__file__))
        # glob.glob is a file search function. It finds all files matching a wildcard pattern 
        # and returns them as a list of paths. 
        # os.path.join is used to join the paths of the directory and the file.
        # * is a wildcard that matches any sequence of characters.
        # TL_data_*.csv will match all files starting with TL_data_ and ending with .csv
        # The list of existing CSVs is stored in the variable existing_csvs.
        existing_csvs = glob.glob(os.path.join(script_dir, "TL_data_*.csv"))
        for old_csv in existing_csvs:          # remove old file(s) before writing
            os.remove(old_csv)

        # Remove old Excel files before writing a new one (same cleanup as CSV)
        existing_xlsxs = glob.glob(os.path.join(script_dir, "TL_data_*.xlsx"))
        for old_xlsx in existing_xlsxs:
            os.remove(old_xlsx)

        csv_path  = os.path.join(script_dir, f"TL_data_{timestamp}.csv")

        try:
            with open(csv_path, "w", newline="") as csv_file:
                writer = csv.writer(csv_file)
                # Write PSN as a metadata row at the top of the CSV
                writer.writerow(["PSN : ", psn_string])
                writer.writerow(["Date/time-stamp", "Pair", "Set_Temp_Dec", "Measured_Temp_Dec"])
                # Write paired temperature set and measured values
                writer.writerows(paired_rows)
            print(f"  [TL] CSV saved → {csv_path}  ({len(paired_rows)} paired rows, {read_errors} errors)")
        except OSError as e:
            print(f"  [TL] Failed to write CSV: {e}")

        # Generate Excel Table and record set/measured temperature values
        create_record_temperatures_excel_table(script_dir, timestamp, paired_rows, psn_string, error_string)

        # Reset status so we don't re-read until next TL_Ready pulse
        read_status_registers.val_busy_ready_STATUS = 0
    
'''
Main program starts here
'''
def main():

    print(" Software Version: 1.0.0.10")
    print("=" * 60)
    print("  PI5 Modbus/RTU Master - PIC MF40 Controller")
    print(f"  Port: {SERIAL_PORT}  Baud: {BAUDRATE}  Slave ID: {SLAVE_ID}")
    print("=" * 60)

    # Create Modbus RTU client (Pi is the MASTER, PIC is the SLAVE)
    client = ModbusSerialClient(
        port=SERIAL_PORT,
        baudrate=BAUDRATE,
        bytesize=BYTESIZE,
        parity=PARITY,
        stopbits=STOPBITS,
        timeout=TIMEOUT,
    )

    try:
        # Connect to serial port
        if not client.connect():
            print(f"Error: Could not connect to {SERIAL_PORT}")
            return
        
        print(f"\nConnected to {SERIAL_PORT} successfully.")

        # Read all registers once
        while(True):
            read_status_registers(client)
            time.sleep(1)
            #read_all_rw_registers(client)
            #read_all_ro_registers(client)
    
    except KeyboardInterrupt:
        print("\n\nProgram stopped by user (Ctrl+C).")

    except serial.SerialException as e:
        print(f"\nSerial port error: {e}")

    except Exception as e:
        print(f"\nUnexpected error: {e}")

    finally:
        client.close()
        print("Modbus connection closed.")


if __name__ == "__main__":

    # Excel workbook for recording temperature data
    workbook = Workbook()
    main()
